import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def save_to_excel(df, output_path):
    """
    Saves the DataFrame to an Excel file with formatting.
    """
    # Write basic data
    df.to_excel(output_path, index=False, sheet_name='WhatsApp Chat')
    
    # Apply formatting
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        try:
            for cell in column:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            # Cap width at 50 to prevent huge columns for long messages
            if adjusted_width > 50:
                adjusted_width = 50
            
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        except:
            pass
            
    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    wb.save(output_path)
